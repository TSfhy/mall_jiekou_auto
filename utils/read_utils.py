"""
文件读取工具类：封装Excel/CSV/YAML/JSON/TXT读取方法
核心：直接接收文件完整路径，支持绝对路径/相对路径，适配更多场景
"""
import os
import json
import csv
# import yaml

from openpyxl import load_workbook
from typing import Union, List, Dict, Any

from config import datas_path


class ReadUtil:
    """文件读取工具类 - 基于完整文件路径读取"""

    @staticmethod
    def _check_file_exists(file_path: str) -> None:
        """
        校验文件是否存在，不存在则抛异常
        :param file_path: 文件完整路径（绝对/相对）
        """
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"文件不存在：{file_path}")
        if not os.path.isfile(file_path):
            raise ValueError(f"不是有效文件：{file_path}")

    @staticmethod
    def read_excel(
        file_path: str,
        sheet_name: str = None,
        header_row: int = 0
    ) -> List[Dict[str, Any]]:
        """
        读取Excel文件（支持.xlsx/.xls），返回列表嵌套字典（表头为key）
        :param file_path: Excel文件完整路径（如：D:/data/login_data.xlsx 或 ./data/login_data.xlsx）
        :param sheet_name: 工作表名（None则读取第一个sheet）
        :param header_row: 表头行号（默认第0行，xlrd从0开始，openpyxl自动适配）
        :return: [{"phone": "13800138000", "pwd": "123456", "expect": True}, ...]
        """
        # 校验文件
        ReadUtil._check_file_exists(file_path)
        # 区分Excel格式
        is_xlsx = file_path.endswith(".xlsx")
        is_xls = file_path.endswith(".xls")
        if not (is_xlsx or is_xls):
            raise ValueError(f"不支持的Excel格式，仅支持.xlsx/.xls：{file_path}")

        # 打开工作簿
        if is_xlsx:
            wb = load_workbook(file_path, data_only=True)  # 读取公式计算后的值
            # 选择工作表
            sheet = wb[sheet_name] if sheet_name else wb.active
            # 提取表头（openpyxl行号从1开始，header_row=0对应第1行）
            # headers = [cell.value for cell in sheet[header_row + 1]]
            # 提取数据行
            data_list = []
            start_row = header_row + 2  # 数据行从表头下一行开始
            # for row in sheet.iter_rows(min_row=start_row, values_only=True):
            #     if all(cell is None for cell in row):  # 跳过空行
            #         continue
            #     row_dict = dict(zip(headers, row))
            #     data_list.append(row_dict)
            for row in sheet.iter_rows(min_row=start_row):

                data = tuple(col.value if col.value is not None else '' for col in row)
                data_list.append(data)

            wb.close()



        return data_list

    @staticmethod
    def read_csv(
        file_path: str,
        delimiter: str = ",",
        encoding: str = "utf-8"
    ) -> List[Dict[str, str]]:
        """
        读取CSV文件，返回列表嵌套字典（第一行为表头）
        :param file_path: CSV文件完整路径
        :param delimiter: 分隔符（默认逗号）
        :param encoding: 编码（默认utf-8，中文乱码试gbk）
        :return: [{"phone": "13800138000", "pwd": "123456"}, ...]
        """
        ReadUtil._check_file_exists(file_path)
        data_list = []
        with open(file_path, "r", encoding=encoding, newline="") as f:
            # 按表头封装为字典
            reader = csv.DictReader(f, delimiter=delimiter)
            for row in reader:
                # 去除值两端空格
                clean_row = {k: v.strip() for k, v in row.items()}
                data_list.append(clean_row)
        return data_list

    # @staticmethod
    # def read_yaml(file_path: str, encoding: str = "utf-8") -> Union[Dict, List]:
    #     """
    #     读取YAML文件，返回字典/列表（适配YAML根节点类型）
    #     :param file_path: YAML文件完整路径
    #     :return: 字典/列表
    #     """
    #     ReadUtil._check_file_exists(file_path)
    #     with open(file_path, "r", encoding=encoding) as f:
    #         # safe_load避免yaml注入风险，更安全
    #         data = yaml.safe_load(f) or {}  # 空文件返回空字典
    #     return data

    @staticmethod
    def read_json(file_path: str, encoding: str = "utf-8") -> Union[Dict, List]:
        """
        读取JSON文件，返回字典/列表
        :param file_path: JSON文件完整路径
        :return: 字典/列表
        """
        ReadUtil._check_file_exists(file_path)
        with open(file_path, "r", encoding=encoding) as f:
            data = json.load(f) or {}
        return data

    @staticmethod
    def read_txt(
        file_path: str,
        encoding: str = "utf-8",
        split: str = None
    ) -> List[Union[str, List[str]]]:
        """
        读取TXT文件，可选按分隔符拆分每行数据
        :param file_path: TXT文件完整路径
        :param split: 分隔符（None返回每行字符串，如","则拆分每行）
        :return: 列表（每行数据）
        """
        ReadUtil._check_file_exists(file_path)
        data_list = []
        with open(file_path, "r", encoding=encoding) as f:
            for line in f:
                line = line.strip()  # 去除换行符/首尾空格
                if not line:  # 跳过空行
                    continue
                if split:
                    # 按分隔符拆分，去除每个元素空格
                    line_data = [item.strip() for item in line.split(split)]
                    data_list.append(line_data)
                else:
                    data_list.append(line)
        return data_list


# 测试示例（可删除，验证功能）
if __name__ == "__main__":
    # 1. 读取Excel（相对路径示例）
    excel_path = os.path.join(datas_path,"login_datas.xlsx")
    excel_data = ReadUtil.read_excel(excel_path, sheet_name="login_data")
    print("Excel读取结果：", excel_data)  # 打印第一条数据
    #
    # # 2. 读取YAML（绝对路径示例，替换为你的实际路径）
    # yaml_path = "D:/app_test_framework/config/app_config.yaml"
    # yaml_data = ReadUtil.read_yaml(yaml_path)
    # print("YAML读取结果：", yaml_data)
    #
    # # 3. 读取CSV
    # csv_path = "./data/csv/login_data.csv"
    # csv_data = ReadUtil.read_csv(csv_path, encoding="gbk")
    # print("CSV读取结果：", csv_data[:1])